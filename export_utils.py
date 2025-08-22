import pandas as pd
import io
import os
from datetime import datetime
from audit_logger import log_audit_event
import tempfile
import json

def generate_export(user_email, user_role, export_format, include_drafts=False, include_rejected=False):
    """Generate export file with enhanced multi-sheet support"""

    if export_format == 'excel_complete':
        # New complete Excel export with all sheets
        from file_handler import export_excel_with_all_sheets
        return export_excel_with_all_sheets(user_email, user_role, include_updates=True)
    elif export_format == 'excel':
        return generate_excel_export_enhanced(user_email, user_role, include_drafts, include_rejected)
    elif export_format == 'csv':
        return generate_csv_export(user_email, user_role, include_drafts, include_rejected)
    elif export_format == 'pdf':
        return generate_pdf_export(user_email, user_role, include_drafts, include_rejected)
    else:
        raise Exception(f"Unsupported export format: {export_format}")

def generate_excel_export_enhanced(user_email, user_role, include_drafts=False, include_rejected=False):
    """Generate enhanced Excel export with original sheets and updates"""
    from models import ExcelFileData, ExcelSheetData, ROPARecord, User
    from app import db

    try:
        user = User.query.filter_by(email=user_email).first()
        if not user:
            raise Exception("User not found")

        temp_dir = tempfile.gettempdir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ROPA_Enhanced_Export_{timestamp}.xlsx"
        file_path = os.path.join(temp_dir, filename)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Export current ROPA records
            records_df = get_filtered_ropa_data(user_email, user_role, include_drafts, include_rejected)
            if not records_df.empty:
                export_df = prepare_export_dataframe(records_df)
                export_df.to_excel(writer, sheet_name='Current_ROPA_Records', index=False)

            # Export original Excel sheets if user has uploaded files
            if user_role == 'Privacy Officer':
                excel_files = ExcelFileData.query.all()
            else:
                excel_files = ExcelFileData.query.filter_by(uploaded_by=user.id).all()

            for excel_file in excel_files:
                sheets = ExcelSheetData.query.filter_by(excel_file_id=excel_file.id).all()
                for sheet in sheets:
                    try:
                        sheet_data = json.loads(sheet.sheet_data)
                        if sheet_data:
                            df = pd.DataFrame(sheet_data)
                            sheet_name = f"Original_{sheet.sheet_name}"[:31]  # Excel limit
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                    except Exception as e:
                        print(f"Error exporting sheet {sheet.sheet_name}: {str(e)}")

            # Add summary sheet
            summary_data = create_enhanced_summary_data(records_df, excel_files)
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Export_Summary', index=False)

        return file_path, filename

    except Exception as e:
        print(f"Error generating enhanced Excel export: {str(e)}")
        raise e

def get_filtered_ropa_data(user_email, user_role, include_drafts=False, include_rejected=False):
    """Get filtered ROPA data for export"""
    from models import ROPARecord, User
    from app import db

    try:
        user = User.query.filter_by(email=user_email).first()
        if not user:
            return pd.DataFrame()

        # Build query based on user role
        if user_role == 'Privacy Champion':
            query = ROPARecord.query.filter_by(created_by=user.id)
        else:
            query = ROPARecord.query

        # Status filtering
        status_filters = ['Approved', 'Under Review']
        if include_drafts:
            status_filters.append('Draft')
        if include_rejected:
            status_filters.append('Rejected')

        # Apply status filter
        records = query.filter(ROPARecord.status.in_(status_filters)).order_by(ROPARecord.created_at.desc()).all()

        # Convert to DataFrame
        if not records:
            return pd.DataFrame()

        records_data = []
        for record in records:
            creator = User.query.get(record.created_by)
            record_dict = {
                'id': record.id,
                'processing_activity_name': record.processing_activity_name,
                'category': record.category,
                'description': record.description,
                'department_function': record.department_function,
                'controller_name': record.controller_name,
                'controller_contact': record.controller_contact,
                'controller_address': record.controller_address,
                'dpo_name': record.dpo_name,
                'dpo_contact': record.dpo_contact,
                'processing_purpose': record.processing_purpose,
                'legal_basis': record.legal_basis,
                'data_categories': record.data_categories,
                'data_subjects': record.data_subjects,
                'recipients': record.recipients,
                'retention_period': record.retention_period,
                'security_measures': record.security_measures,
                'status': record.status,
                'created_by': creator.email if creator else 'Unknown',
                'created_at': record.created_at,
                'updated_at': record.updated_at
            }
            records_data.append(record_dict)

        return pd.DataFrame(records_data)

    except Exception as e:
        print(f"Error getting filtered ROPA data: {str(e)}")
        return pd.DataFrame()

def generate_csv_export(user_email, user_role, include_drafts=False, include_rejected=False):
    """Generate CSV export"""
    records_df = get_filtered_ropa_data(user_email, user_role, include_drafts, include_rejected)

    if records_df.empty:
        raise Exception("No records found for export")

    export_df = prepare_export_dataframe(records_df)

    temp_dir = tempfile.gettempdir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ROPA_Export_{timestamp}.csv"
    file_path = os.path.join(temp_dir, filename)

    export_df.to_csv(file_path, index=False)
    return file_path, filename

def generate_pdf_export(user_email, user_role, include_drafts=False, include_rejected=False):
    """Generate PDF report"""
    records_df = get_filtered_ropa_data(user_email, user_role, include_drafts, include_rejected)

    report_content = create_pdf_report_content(records_df)

    temp_dir = tempfile.gettempdir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ROPA_Report_{timestamp}.txt"
    file_path = os.path.join(temp_dir, filename)

    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(report_content)

    return file_path, filename

def prepare_export_dataframe(df):
    """Prepare dataframe for export with clean column names"""
    if df.empty:
        return df

    export_columns = {
        'processing_activity_name': 'Processing Activity Name',
        'category': 'Category',
        'description': 'Description',
        'department_function': 'Department/Function',
        'controller_name': 'Controller Name',
        'controller_contact': 'Controller Contact',
        'controller_address': 'Controller Address',
        'dpo_name': 'DPO Name',
        'dpo_contact': 'DPO Contact',
        'processing_purpose': 'Purpose of Processing',
        'legal_basis': 'Legal Basis',
        'data_categories': 'Data Categories',
        'data_subjects': 'Data Subjects',
        'recipients': 'Recipients',
        'retention_period': 'Retention Period',
        'security_measures': 'Security Measures',
        'status': 'Status',
        'created_by': 'Created By',
        'created_at': 'Created Date',
        'updated_at': 'Updated Date'
    }

    # Select available columns
    available_columns = {k: v for k, v in export_columns.items() if k in df.columns}

    export_df = df[list(available_columns.keys())].copy()
    export_df = export_df.rename(columns=available_columns)

    # Clean up date columns
    date_columns = ['Created Date', 'Updated Date']
    for col in date_columns:
        if col in export_df.columns:
            export_df[col] = pd.to_datetime(export_df[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S')

    return export_df

def create_enhanced_summary_data(records_df, excel_files):
    """Create enhanced summary with original file information"""
    summary = []

    # Basic statistics
    summary.append({"Metric": "Total ROPA Records", "Value": len(records_df)})
    if not records_df.empty:
        summary.append({"Metric": "Approved Records", "Value": len(records_df[records_df['status'] == 'Approved'])})
        summary.append({"Metric": "Pending Review", "Value": len(records_df[records_df['status'] == 'Under Review'])})
        summary.append({"Metric": "Draft Records", "Value": len(records_df[records_df['status'] == 'Draft'])})

    # Excel file information
    summary.append({"Metric": "Original Excel Files", "Value": len(excel_files)})

    for excel_file in excel_files:
        summary.append({"Metric": f"File: {excel_file.filename}", "Value": f"Uploaded: {excel_file.upload_timestamp.strftime('%Y-%m-%d %H:%M')}"})
        summary.append({"Metric": f"Sheets in {excel_file.filename}", "Value": excel_file.total_sheets})

    summary.append({"Metric": "Export Generated", "Value": datetime.now().strftime('%Y-%m-%d %H:%M:%S')})

    return summary

def create_pdf_report_content(df):
    """Create text content for PDF report"""
    report = f"""
GDPR ROPA COMPLIANCE REPORT
Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

========================================
EXECUTIVE SUMMARY
========================================

Total ROPA Records: {len(df)}
"""

    if not df.empty:
        report += f"""Approved Records: {len(df[df['status'] == 'Approved'])}
Pending Review: {len(df[df['status'] == 'Under Review'])}
Draft Records: {len(df[df['status'] == 'Draft'])}

Compliance Rate: {(len(df[df['status'] == 'Approved']) / len(df) * 100):.1f}%
"""

    report += """
========================================
RECORDS BY CATEGORY
========================================
"""

    if not df.empty and 'category' in df.columns:
        category_counts = df['category'].value_counts()
        for category, count in category_counts.items():
            report += f"{category}: {count} records\n"

    return report


def export_excel_with_all_sheets(user_email, user_role, include_updates=True):
    """Export Excel file with all original sheets plus updates - beautifully formatted"""
    from models import ExcelFileData, ExcelSheetData, ROPARecord, User
    from app import db
    import tempfile
    import os
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook

    try:
        # Get user's uploaded files
        user = User.query.filter_by(email=user_email).first()
        if not user:
            raise Exception("User not found")

        # Get the most recent Excel file uploaded by user or all files for Privacy Officer
        if user_role == 'Privacy Officer':
            excel_files = ExcelFileData.query.order_by(ExcelFileData.upload_timestamp.desc()).all()
        else:
            excel_files = ExcelFileData.query.filter_by(uploaded_by=user.id).order_by(ExcelFileData.upload_timestamp.desc()).all()

        if not excel_files:
            raise Exception("No Excel files found to export")

        # Create new Excel file with all sheets
        temp_dir = tempfile.gettempdir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ROPA_Export_{timestamp}.xlsx"
        file_path = os.path.join(temp_dir, filename)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            sheets_written = 0
            workbook = writer.book

            # Process each uploaded Excel file
            for excel_file in excel_files:
                # Get all sheets for this file, ordered by original sheet order
                sheets = ExcelSheetData.query.filter_by(excel_file_id=excel_file.id).all()

                # Sort sheets to maintain original order if possible
                original_sheet_names = json.loads(excel_file.sheet_names)
                sheets_dict = {sheet.sheet_name: sheet for sheet in sheets}

                for original_name in original_sheet_names:
                    if original_name in sheets_dict:
                        sheet = sheets_dict[original_name]
                        try:
                            # Load original sheet data
                            sheet_data = json.loads(sheet.sheet_data)

                            if sheet_data:
                                # Convert to DataFrame while preserving EXACT original structure
                                if sheet_data and len(sheet_data) > 0:
                                    # Create DataFrame from the stored data
                                    df = pd.DataFrame(sheet_data)
                                    
                                    # The data is already stored with the original column names
                                    # Don't modify column names at all - keep exactly as uploaded
                                    # The sheet_data already contains the data with original headers preserved
                                    
                                    # Remove any completely empty rows but keep original column structure
                                    df = df.dropna(how='all')
                                else:
                                    df = pd.DataFrame()

                                # Clean sheet name for Excel compatibility
                                sheet_name = str(original_name).strip()
                                
                                # Remove any problematic characters but keep the original name
                                invalid_chars = ['[', ']', '*', '?', ':', '/', '\\']
                                for char in invalid_chars:
                                    sheet_name = sheet_name.replace(char, '_')

                                # Handle Excel sheet name length limit (31 chars)
                                if len(sheet_name) > 31:
                                    sheet_name = sheet_name[:31].rstrip('_')

                                # Ensure unique sheet names if multiple files
                                original_sheet_name = sheet_name
                                counter = 1
                                while sheet_name in [ws.title for ws in workbook.worksheets]:
                                    suffix = f"_{counter}"
                                    max_base_length = 31 - len(suffix)
                                    sheet_name = original_sheet_name[:max_base_length] + suffix
                                    counter += 1

                                # Write sheet with formatting - ensure sheet_name is valid
                                if not sheet_name or sheet_name.isspace():
                                    sheet_name = f"Sheet_{sheets_written + 1}"
                                
                                # Use header=False to preserve exact structure
                                df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
                                worksheet = workbook[sheet_name]

                                # Apply beautiful formatting with colors
                                format_excel_sheet(worksheet, df, is_original_sheet=True, original_sheet_name=original_name)
                                sheets_written += 1

                                print(f"Exported sheet: '{original_name}' as '{sheet_name}' with {len(df)} rows and {len(df.columns)} columns")

                        except Exception as e:
                            print(f"Error writing sheet {original_name}: {str(e)}")
                            # Try with a simple fallback name
                            try:
                                fallback_name = f"Sheet_{sheets_written + 1}"
                                if sheet_data:
                                    df = pd.DataFrame(sheet_data)
                                    df.to_excel(writer, sheet_name=fallback_name, index=False)
                                    worksheet = workbook[fallback_name]
                                    format_excel_sheet(worksheet, df, is_original_sheet=True)
                                    sheets_written += 1
                                    print(f"Used fallback name: '{fallback_name}' for sheet '{original_name}'")
                            except Exception as fallback_error:
                                print(f"Fallback also failed for {original_name}: {str(fallback_error)}")

            # Integrate approved custom fields into original sheets
            try:
                from custom_tab_automation import get_approved_custom_fields_by_category
                custom_fields = get_approved_custom_fields_by_category()
                
                # Get custom field data for integration
                custom_field_values = {}
                if custom_fields:
                    from models import ROPACustomData, ROPARecord
                    for category, fields in custom_fields.items():
                        for field in fields:
                            field_values = db.session.query(ROPACustomData, ROPARecord).join(
                                ROPARecord, ROPACustomData.ropa_record_id == ROPARecord.id
                            ).filter(ROPACustomData.custom_field_id == field['id']).all()
                            
                            for custom_data, ropa_record in field_values:
                                if user_role != 'Privacy Officer' and ropa_record.created_by != user.id:
                                    continue
                                
                                record_key = ropa_record.processing_activity_name or f"Record_{ropa_record.id}"
                                if record_key not in custom_field_values:
                                    custom_field_values[record_key] = {}
                                
                                custom_field_values[record_key][field['field_name']] = custom_data.field_value or ''
                
                # Now enhance existing sheets with custom field columns
                if custom_field_values:
                    enhance_existing_sheets_with_custom_fields(workbook, custom_field_values, custom_fields)
                            
            except Exception as e:
                print(f"Error integrating custom fields: {str(e)}")

            # Custom field integration is handled in enhance_existing_sheets_with_custom_fields

            # Add export summary sheet
            try:
                summary_data = create_export_summary(excel_files, [])
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Export_Summary', index=False)

                # Format summary sheet
                worksheet = workbook['Export_Summary']
                format_summary_sheet(worksheet, summary_df)
                sheets_written += 1

            except Exception as e:
                print(f"Error creating summary sheet: {str(e)}")

        return file_path, filename

    except Exception as e:
        print(f"Error exporting Excel with all sheets: {str(e)}")
        raise e

def format_excel_sheet(worksheet, df, is_ropa_sheet=False, is_original_sheet=False, original_sheet_name=None):
    """Apply professional formatting to Excel sheet with enhanced visual appeal"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Professional color schemes matching uploaded ROPA style
        if is_original_sheet:
            header_color = "2F4F7F"  # Professional dark blue for original sheets
            alt_row_color = "E8F1FF"  # Very light blue
            accent_color = "B4D7FF"  # Medium blue for emphasis
        elif is_ropa_sheet:
            header_color = "1F497D"  # Deep blue for ROPA sheets  
            alt_row_color = "F0F4F8"  # Light gray-blue
            accent_color = "DDEEFF"  # Light blue accent
        else:
            header_color = "4F6228"  # Professional dark green
            alt_row_color = "F2F8E8"  # Light green
            accent_color = "E8F4D0"  # Medium green

        # Enhanced header formatting with professional styling
        header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_border = Border(
            left=Side(style='medium', color="FFFFFF"),
            right=Side(style='medium', color="FFFFFF"),
            top=Side(style='medium', color="FFFFFF"),
            bottom=Side(style='medium', color="FFFFFF")
        )

        # Ensure we have actual data to format
        if df.empty:
            return

        # Apply professional header formatting
        num_cols = len(df.columns)
        for col_num in range(1, num_cols + 1):
            cell = worksheet.cell(row=1, column=col_num)
            
            # Set header value with proper formatting
            if col_num <= len(df.columns):
                header_value = str(df.columns[col_num - 1])
                
                # If header is empty or unnamed, provide a meaningful default
                if not header_value or header_value.strip() == '' or header_value.lower() in ['unnamed', 'nan', 'none']:
                    header_value = f"Column {col_num}"
                
                # Clean up header text while preserving original meaning
                header_value = header_value.strip()
                if len(header_value) > 50:  # Truncate very long headers
                    header_value = header_value[:47] + "..."
                
                cell.value = header_value
            
            # Apply professional header styling
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

        # Enhanced data cell formatting with professional styling
        data_font = Font(name="Calibri", size=11, color="2C2C2C")
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        data_border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC")
        )

        # Apply data formatting with professional alternating colors
        num_rows = len(df) + 1  # +1 for header
        for row_num in range(2, num_rows + 1):
            for col_num in range(1, num_cols + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                
                # Ensure cell has a value (handle None/NaN)
                if cell.value is None or str(cell.value).lower() in ['nan', 'none']:
                    cell.value = ''
                
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border

                # Professional alternating row coloring with better contrast
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color=alt_row_color, end_color=alt_row_color, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Intelligent column width calculation for better readability
        for col_num in range(1, num_cols + 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate optimal width based on content
            header_length = len(str(worksheet.cell(row=1, column=col_num).value or ''))
            max_length = max(header_length, 15)  # Minimum readable width
            
            # Sample data to determine optimal width (check more rows for accuracy)
            sample_size = min(num_rows, 100)  # Check up to 100 rows
            for row_num in range(2, sample_size + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    try:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                    except:
                        pass

            # Set optimal column width with professional limits
            if max_length <= 20:
                adjusted_width = max_length + 3
            elif max_length <= 40:
                adjusted_width = max_length + 2
            else:
                adjusted_width = 45  # Cap very wide columns
            
            # Ensure minimum and maximum bounds
            adjusted_width = min(max(adjusted_width, 12), 55)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Professional row height settings
        worksheet.row_dimensions[1].height = 45  # Generous header height for readability

        # Set comfortable row height for data rows
        for row_num in range(2, min(num_rows + 1, 500)):  # Format up to 500 rows
            worksheet.row_dimensions[row_num].height = 30  # Comfortable reading height

        # Enable gridlines for better structure visibility
        worksheet.sheet_view.showGridLines = True
        
        # Freeze the header row for better navigation
        worksheet.freeze_panes = 'A2'

        print(f"Applied professional formatting to sheet '{original_sheet_name or 'Unknown'}' with {num_cols} columns and {num_rows-1} data rows")

    except Exception as e:
        print(f"Error formatting Excel sheet {original_sheet_name or 'Unknown'}: {str(e)}")
        # Enhanced fallback formatting
        try:
            if not df.empty:
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
                    cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium')
                    )
                    
                    # Set column width
                    column_letter = get_column_letter(col_num)
                    worksheet.column_dimensions[column_letter].width = 20
                    
                # Set row heights
                worksheet.row_dimensions[1].height = 40
                for row in range(2, min(len(df) + 2, 100)):
                    worksheet.row_dimensions[row].height = 25
                    
        except Exception as fallback_error:
            print(f"Even fallback formatting failed: {str(fallback_error)}")

def format_summary_sheet(worksheet, df):
    """Format the summary sheet with special styling"""
    try:
        # Title formatting
        title_font = Font(bold=True, color="FFFFFF", size=14)
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")

        # Merge cells for title
        worksheet.merge_cells('A1:B1')
        title_cell = worksheet['A1']
        title_cell.value = "ROPA Export Summary"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment

        # Shift data down by 2 rows
        for row_num in range(len(df), 0, -1):
            for col_num in range(1, len(df.columns) + 1):
                old_cell = worksheet.cell(row=row_num + 1, column=col_num)
                new_cell = worksheet.cell(row=row_num + 3, column=col_num)
                new_cell.value = old_cell.value

        # Clear the old data
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                worksheet.cell(row=row_num, column=col_num).value = None

        # Format the summary data using the general format_excel_sheet for consistency
        # Adjust the number of columns for formatting based on the actual data in df
        num_cols_to_format = len(df.columns)
        if num_cols_to_format > 0:
            format_excel_sheet(worksheet, df) # This will format up to df's columns

    except Exception as e:
        print(f"Error formatting summary sheet: {str(e)}")

def create_export_summary(excel_files, ropa_records):
    """Create export summary data"""
    summary = []

    # Export information
    summary.append({"Category": "Export Details", "Information": f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"})
    summary.append({"Category": "Total Original Files", "Information": str(len(excel_files))})

    # File details
    for idx, excel_file in enumerate(excel_files, 1):
        summary.append({"Category": f"File {idx}", "Information": excel_file.filename})
        summary.append({"Category": f"File {idx} Upload Date", "Information": excel_file.upload_timestamp.strftime('%Y-%m-%d %H:%M:%S')})
        summary.append({"Category": f"File {idx} Sheets", "Information": str(excel_file.total_sheets)})

        # List sheet names
        sheet_names = json.loads(excel_file.sheet_names)
        summary.append({"Category": f"File {idx} Sheet Names", "Information": ", ".join(sheet_names)})

    # ROPA records information
    summary.append({"Category": "System ROPA Records", "Information": str(len(ropa_records))})

    if ropa_records:
        status_counts = {}
        for record in ropa_records:
            status = record.status or 'Unknown'
            status_counts[status] = status_counts.get(status, 0) + 1

        for status, count in status_counts.items():
            summary.append({"Category": f"Records - {status}", "Information": str(count)})

    # Custom fields information
    try:
        from custom_tab_automation import get_approved_custom_fields_by_category
        custom_fields = get_approved_custom_fields_by_category()
        total_custom_fields = sum(len(fields) for fields in custom_fields.values())
        summary.append({"Category": "Approved Custom Fields", "Information": str(total_custom_fields)})
        
        for category, fields in custom_fields.items():
            if fields:
                summary.append({"Category": f"Custom Fields - {category}", "Information": str(len(fields))})
    except Exception as e:
        summary.append({"Category": "Custom Fields", "Information": f"Error retrieving: {str(e)}"})

    return summary


def enhance_existing_sheets_with_custom_fields(workbook, custom_field_values, custom_fields):
    """Enhance existing sheets by adding custom field columns at specific positions"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Get all approved custom field names
        all_custom_fields = []
        for category, fields in custom_fields.items():
            for field in fields:
                all_custom_fields.append(field['field_name'])
        
        if not all_custom_fields:
            return
        
        # Process each worksheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Skip summary and system sheets
            if sheet_name in ['Export_Summary', 'System_ROPA_Records']:
                continue
            
            # Find the insertion position based on sheet type
            insert_position = None
            notes_column = None
            max_col = 1
            
            # Get all header values to find the right position
            headers = []
            for col in range(1, worksheet.max_column + 50):  # Check more columns
                cell_value = worksheet.cell(row=1, col=col).value
                if cell_value:
                    headers.append((col, str(cell_value).strip().lower()))
                    max_col = col
                elif col <= max_col + 10:  # Check a few more columns after last data
                    headers.append((col, ''))
                else:
                    break
            
            # Find Notes/Comments column first
            for col_num, header_text in headers:
                if any(keyword in header_text for keyword in ['notes', 'comments', 'note', 'comment']):
                    notes_column = col_num
                    insert_position = col_num  # Insert before Notes/Comments
                    break
            
            # If Notes/Comments not found, determine insertion position based on sheet content
            if insert_position is None:
                for col_num, header_text in headers:
                    # For Controller sheets - insert after "Reasons For Not Adhering to Policy"
                    if any(keyword in header_text for keyword in ['reasons for not adhering', 'reasons for not adhering to policy', 'policy adherence']):
                        insert_position = col_num + 1
                        break
                    # For Processor sheets - insert after "Safeguards & Measures"  
                    elif any(keyword in header_text for keyword in ['safeguards', 'safeguards & measures', 'safeguards and measures']):
                        insert_position = col_num + 1
                        break
                
                # If still not found, insert at the end
                if insert_position is None:
                    insert_position = max_col + 1
            
            # If Notes/Comments column exists, move it to after custom fields
            if notes_column is not None:
                # Move the Notes/Comments column to after custom fields
                for row_num in range(1, worksheet.max_row + 1):
                    old_cell = worksheet.cell(row=row_num, column=notes_column)
                    new_cell = worksheet.cell(row=row_num, column=insert_position + len(all_custom_fields))
                    new_cell.value = old_cell.value
                    new_cell.font = old_cell.font
                    new_cell.fill = old_cell.fill
                    new_cell.alignment = old_cell.alignment
                    new_cell.border = old_cell.border
                    # Clear the old cell
                    old_cell.value = None
                    old_cell.font = Font()
                    old_cell.fill = PatternFill()
                    old_cell.alignment = Alignment()
                    old_cell.border = Border()
            
            # Add custom field headers
            for idx, field_name in enumerate(all_custom_fields):
                col_num = insert_position + idx
                
                # Set header styling to match existing headers
                header_cell = worksheet.cell(row=1, column=col_num)
                header_cell.value = field_name
                header_cell.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
                header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                header_cell.border = Border(
                    left=Side(style='thick', color="FFFFFF"),
                    right=Side(style='thick', color="FFFFFF"),
                    top=Side(style='thick', color="FFFFFF"),
                    bottom=Side(style='thick', color="FFFFFF")
                )
                
                # If there's a second header row, add field name there too
                if worksheet.max_row >= 2 and worksheet.cell(row=2, column=1).value:
                    subheader_cell = worksheet.cell(row=2, column=col_num)
                    subheader_cell.value = field_name
                    subheader_cell.font = Font(bold=True, color="FFFFFF", size=9)
                    subheader_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    subheader_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    subheader_cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Set column width
                column_letter = get_column_letter(col_num)
                worksheet.column_dimensions[column_letter].width = 25
            
            # Fill in custom field data for existing rows
            data_start_row = 3 if worksheet.max_row >= 2 and worksheet.cell(row=2, column=1).value else 2
            
            for row_num in range(data_start_row, worksheet.max_row + 1):
                # Try to identify the record by looking at the first few columns for activity name
                record_identifier = None
                for col in range(1, min(5, max_col)):  # Check first few columns
                    cell_value = worksheet.cell(row=row_num, column=col).value
                    if cell_value and str(cell_value).strip():
                        # Try to match with custom field values
                        if str(cell_value).strip() in custom_field_values:
                            record_identifier = str(cell_value).strip()
                            break
                
                # Fill custom field values
                for idx, field_name in enumerate(all_custom_fields):
                    col_num = insert_position + idx
                    data_cell = worksheet.cell(row=row_num, column=col_num)
                    
                    # Set default styling
                    data_cell.font = Font(name="Arial", size=10, color="333333")
                    data_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    data_cell.border = Border(
                        left=Side(style='thin', color="CCCCCC"),
                        right=Side(style='thin', color="CCCCCC"),
                        top=Side(style='thin', color="CCCCCC"),
                        bottom=Side(style='thin', color="CCCCCC")
                    )
                    
                    # Alternating row colors
                    if row_num % 2 == 0:
                        data_cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
                    else:
                        data_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    
                    # Set the value if we found a matching record
                    if record_identifier and record_identifier in custom_field_values:
                        if field_name in custom_field_values[record_identifier]:
                            data_cell.value = custom_field_values[record_identifier][field_name]
                        else:
                            data_cell.value = ""
                    else:
                        data_cell.value = ""
            
            print(f"Enhanced sheet '{sheet_name}' with {len(all_custom_fields)} custom field columns")
    
    except Exception as e:
        print(f"Error enhancing sheets with custom fields: {str(e)}")
