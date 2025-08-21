
import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import openpyxl

def get_all_database_columns():
    """Get all columns from the ROPARecord table dynamically"""
    try:
        from app import db
        import sqlalchemy

        # Use inspector to get actual database columns (including dynamically added ones)
        inspector = sqlalchemy.inspect(db.engine)
        table_columns = inspector.get_columns('ropa_records')
        columns = [col['name'] for col in table_columns]

        # Filter out system columns
        excluded_columns = ['id', 'created_by', 'reviewed_by']
        return [col for col in columns if col not in excluded_columns]
    except Exception as e:
        print(f"Error getting database columns: {str(e)}")
        # Fallback to model columns if inspector fails
        try:
            from models import ROPARecord
            columns = [column.name for column in ROPARecord.__table__.columns]
            excluded_columns = ['id', 'created_by', 'reviewed_by']
            return [col for col in columns if col not in excluded_columns]
        except:
            return []

def get_all_ropa_data_for_template():
    """Get all existing ROPA records as a pandas DataFrame for template population"""
    try:
        from database import get_db_connection
        import sqlite3

        conn = get_db_connection()
        cursor = conn.cursor()
        
        # First, let's check what columns actually exist in the database
        cursor.execute("PRAGMA table_info(ropa_records)")
        table_info = cursor.fetchall()
        actual_columns = [col[1] for col in table_info]
        print(f"Actual database columns: {actual_columns}")
        
        # Get all records with all available columns
        query = "SELECT * FROM ropa_records ORDER BY created_at DESC"
        
        cursor.execute(query)
        rows = cursor.fetchall()
        
        # Get column names from cursor description
        columns = [description[0] for description in cursor.description]
        
        # Create DataFrame
        df = pd.DataFrame(rows, columns=columns)
        conn.close()
        
        print(f"Found {len(df)} ROPA records for template")
        print(f"Retrieved columns: {list(df.columns)}")
        
        if df.empty:
            return pd.DataFrame()

        # Debug: Print actual values from first record before any processing
        if len(df) > 0:
            print("RAW first record data from database:")
            first_record = df.iloc[0]
            for col in ['id', 'processing_activity_name', 'controller_name', 'controller_contact', 'controller_address', 'dpo_name']:
                if col in df.columns:
                    raw_value = first_record[col]
                    print(f"  {col}: '{raw_value}' (type: {type(raw_value)})")
        
        # Clean up the data more carefully
        for col in df.columns:
            if col in ['created_at', 'updated_at', 'reviewed_at', 'approved_at']:
                # Keep datetime columns as strings if they're already formatted
                df[col] = df[col].fillna('')
            elif col == 'dpia_required':
                # Convert boolean/int to Yes/No
                df[col] = df[col].apply(lambda x: 'Yes' if x in [1, True, 'Yes', 'yes'] else 'No' if x in [0, False, 'No', 'no'] else str(x) if pd.notna(x) and str(x).strip() else '')
            else:
                # Handle other columns - be more careful with None/null values
                df[col] = df[col].apply(lambda x: str(x).strip() if x is not None and pd.notna(x) and str(x).strip() not in ['nan', 'None', 'NaT', 'NULL'] else '')

        print(f"DataFrame created with {len(df)} rows and {len(df.columns)} columns")
        
        # Debug: Show cleaned data for first record
        if len(df) > 0:
            print("First record after cleaning - key fields:")
            first_record = df.iloc[0]
            key_fields = ['id', 'processing_activity_name', 'controller_name', 'controller_contact', 'controller_address', 'dpo_name', 'processing_purpose', 'status']
            for col in key_fields:
                if col in df.columns:
                    value = first_record[col]
                    print(f"  {col}: '{value}'")
        
        return df

    except Exception as e:
        print(f"Error getting all ROPA data for template: {str(e)}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()

def read_uploaded_excel_structure(file_path):
    """Read the uploaded ROPA Excel file to understand its exact structure"""
    try:
        # Load the workbook to get all sheet names
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        print(f"Excel file contains sheets: {sheet_names}")
        
        sheets_data = {}
        for sheet_name in sheet_names:
            try:
                # Try to read each sheet
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                print(f"Sheet '{sheet_name}' structure:")
                print(f"Shape: {df.shape}")
                sheets_data[sheet_name] = df
            except Exception as e:
                print(f"Error reading sheet '{sheet_name}': {str(e)}")
                
        wb.close()
        return sheets_data
    except Exception as e:
        print(f"Error reading Excel file: {str(e)}")
        return None

def create_controller_sheet(wb, existing_data):
    """Create the Controller Processing Activities Register sheet"""
    ws = wb.create_sheet("Controller Processing Activities Register")
    
    # Define the exact column structure from the uploaded ROPA file
    columns_structure = [
        # Controller Details section (3 columns)
        ("Name", "controller_name"),
        ("Address", "controller_address"), 
        ("Contact Details", "controller_contact"),
        
        # Data Protection Officer section (3 columns)
        ("Name", "dpo_name"),
        ("Address", "dpo_address"),
        ("Contact Details", "dpo_contact"),
        
        # Representative Details section (3 columns)
        ("Name", "representative_name"),
        ("Address", "representative_address"),
        ("Contact Details", "representative_contact"),
        
        # Processing Details - map to actual database fields
        ("Detail the department or function responsible for the processing", "department_function"),
        ("Describe the purpose of the processing", "processing_purpose"),
        ("Describe the categories of data subjects", "data_subjects"),
        ("Describe the categories of personal data", "data_categories"),
        ("If possible, envisaged time limits for erasure of different categories of data", "retention_period"),
        ("Categories of recipients to whom personal data has/will be disclosed", "recipients"),
        ("If possible, description of technical & organisational security measures (e.g. pseudonymisation, encryption, codes of conduct etc)", "security_measures"),
        ("Legal Basis for Processing", "legal_basis"),
        ("Is Data Protection Impact Assessment (DPIA) Required?", "dpia_required"),
        ("International transfers", "international_transfers"),
        ("Any other relevant information", "additional_info")
    ]

    # Create the header structure exactly as shown in the image
    # Row 1: Main section headers
    ws.merge_cells('A1:C1')  # Controller Details
    ws['A1'] = 'Controller Details'
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=11)
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('D1:F1')  # Data Protection Officer
    ws['D1'] = 'Data Protection Officer'
    ws['D1'].font = Font(bold=True, color="FFFFFF", size=11)
    ws['D1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('G1:I1')  # Representative Details
    ws['G1'] = 'Representative Details (if applicable)'
    ws['G1'].font = Font(bold=True, color="FFFFFF", size=11)
    ws['G1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['G1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['G1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Calculate the last column letter for Processing Details
    last_col = get_column_letter(len(columns_structure))
    ws.merge_cells(f'J1:{last_col}1')  # Processing Details
    ws['J1'] = 'Processing Details'
    ws['J1'].font = Font(bold=True, color="FFFFFF", size=11)
    ws['J1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['J1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['J1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Row 2: Sub-headers exactly as shown in the image
    headers = [col[0] for col in columns_structure]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Set row heights to match the image
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 80  # Taller for the detailed headers

    # Set column widths to match the proportions in the image
    column_widths = [20, 30, 25, 20, 30, 25, 20, 30, 25, 35, 35, 30, 35, 25, 30, 40, 25, 15, 25, 30]
    for i, width in enumerate(column_widths[:len(columns_structure)], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Add existing database data starting from row 3 (no sample data)
    start_row = 3
    if not existing_data.empty:
        print(f"Populating controller sheet with {len(existing_data)} existing records")
        print(f"Available columns in data: {list(existing_data.columns)}")
        
        for row_idx, (_, row) in enumerate(existing_data.iterrows(), start_row):
            print(f"Processing row {row_idx - start_row + 1}")
            for col_idx, (header, db_field) in enumerate(columns_structure, 1):
                # Get value from existing data, with better field mapping
                value = ''
                
                # Try direct field mapping first
                if db_field and db_field in row:
                    value = row[db_field]
                # Handle special cases and alternative field names
                elif db_field == 'international_transfers':
                    # Try alternative field names
                    value = row.get('third_country_transfers', '') or row.get('international_transfers', '')
                elif db_field == 'additional_info':
                    # Try alternative field names
                    value = row.get('additional_info', '') or row.get('other_relevant_information', '') or row.get('any_other_relevant_information', '')
                elif db_field == 'dpia_required':
                    # Handle boolean DPIA field
                    dpia_val = row.get('dpia_required', '')
                    if isinstance(dpia_val, bool):
                        value = 'Yes' if dpia_val else 'No'
                    elif str(dpia_val).lower() in ['true', '1', 'yes']:
                        value = 'Yes'
                    elif str(dpia_val).lower() in ['false', '0', 'no']:
                        value = 'No'
                    else:
                        value = str(dpia_val) if dpia_val else ''
                
                # Handle None values and convert to string
                if value is None or str(value).lower() in ['nan', 'none']:
                    value = ''
                else:
                    value = str(value).strip()

                # Debug print for first record
                if row_idx == start_row:
                    print(f"  {header} ({db_field}): '{value}'")

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.font = Font(name="Calibri", size=10)

                # Apply alternating row colors
                fill_color = "FFFFFF" if row_idx % 2 == 1 else "F2F2F2"
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            ws.row_dimensions[row_idx].height = 30

        # Add empty rows for data entry after existing data
        next_empty_row = start_row + len(existing_data)
    else:
        next_empty_row = start_row

    # Add empty rows for future data entry
    for row_idx in range(next_empty_row, next_empty_row + 20):
        fill_color = "FFFFFF" if row_idx % 2 == 1 else "F2F2F2"
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_idx in range(1, len(columns_structure) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.fill = fill
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row_idx].height = 30

def create_processor_sheet(wb):
    """Create the Processor Processing Activity sheet exactly as shown in the uploaded file"""
    ws = wb.create_sheet("Processor Processing Activity")
    
    # Define the processor column structure
    processor_columns = [
        # Processor Details section (3 columns)
        ("Name", "processor_name"),
        ("Address", "processor_address"), 
        ("Contact Details", "processor_contact"),
        
        # Data Protection Officer section (3 columns)
        ("Name", "processor_dpo_name"),
        ("Address", "processor_dpo_address"),
        ("Contact Details", "processor_dpo_contact"),
        
        # Representative Details section (3 columns)
        ("Name", "processor_representative_name"),
        ("Address", "processor_representative_address"),
        ("Contact Details", "processor_representative_contact"),
        
        # Processing Details for Processor
        ("Detail the department or function responsible for the processing", "processor_department_function"),
        ("Describe the purpose of the processing", "processor_processing_purpose"),
        ("Describe the categories of data subjects", "processor_data_subjects"),
        ("Describe the categories of personal data", "processor_data_categories"),
        ("If possible, envisaged time limits for erasure of different categories of data", "processor_retention_period"),
        ("Categories of recipients to whom personal data has/will be disclosed", "processor_recipients"),
        ("If possible, description of technical & organisational security measures (e.g. pseudonymisation, encryption, codes of conduct etc)", "processor_security_measures"),
        ("Legal Basis for Processing", "processor_legal_basis"),
        ("Is Data Protection Impact Assessment (DPIA) Required?", "processor_dpia_required"),
        ("International transfers", "processor_international_transfers"),
        ("Any other relevant information", "processor_additional_info")
    ]

    # Create the header structure for Processor sheet
    # Row 1: Main section headers
    ws.merge_cells('A1:C1')  # Processor Details
    ws['A1'] = 'Processor Details'
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=11)
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('D1:F1')  # Data Protection Officer
    ws['D1'] = 'Data Protection Officer'
    ws['D1'].font = Font(bold=True, color="FFFFFF", size=11)
    ws['D1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('G1:I1')  # Representative Details
    ws['G1'] = 'Representative Details (if applicable)'
    ws['G1'].font = Font(bold=True, color="FFFFFF", size=11)
    ws['G1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['G1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['G1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Calculate the last column letter for Processing Details
    last_col = get_column_letter(len(processor_columns))
    ws.merge_cells(f'J1:{last_col}1')  # Processing Details
    ws['J1'] = 'Processing Details'
    ws['J1'].font = Font(bold=True, color="FFFFFF", size=11)
    ws['J1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['J1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['J1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Row 2: Sub-headers
    headers = [col[0] for col in processor_columns]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 80

    # Set column widths
    column_widths = [20, 30, 25, 20, 30, 25, 20, 30, 25, 35, 35, 30, 35, 25, 30, 40, 25, 15, 25, 30]
    for i, width in enumerate(column_widths[:len(processor_columns)], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Add empty rows for future data entry
    for row_idx in range(3, 23):
        fill_color = "FFFFFF" if row_idx % 2 == 1 else "F2F2F2"
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_idx in range(1, len(processor_columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.fill = fill
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row_idx].height = 30

def create_introduction_sheet(wb):
    """Create the Introduction sheet with GDPR ROPA information"""
    ws = wb.create_sheet("Introduction")
    
    # Add title
    ws.merge_cells('A1:I1')
    ws['A1'] = 'PROCESSING ACTIVITIES REGISTER'
    ws['A1'].font = Font(bold=True, size=16, color="000000")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    ws['A1'].border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws.row_dimensions[1].height = 30
    
    # ABOUT THE TEMPLATE section
    ws['A3'] = 'ABOUT THE TEMPLATE'
    ws['A3'].font = Font(bold=True, size=12)
    ws.row_dimensions[3].height = 20
    
    about_text = """Section 10 (1) (2) (a) of the Cyber and Data Protection Regulations [SI 155 of 2024] requires that a data controller shall notify the Authority of all its processing activities carried out on personal information. As you also process some EU citizens data, Article 30 of the GDPR requires that certain organisations maintain records of their processing activities. The specific requirements are noted in our GDPR Policy & Procedure document, and are a mandatory requirement for firms with more than 250 employees and also in certain instances, for firms with less than 250 employees.

The aim of keeping a record of the processing activities is to document the purposes of processing, describe the categories involved, detail disclosures and transfers, and note any time limits for erasing the personal data."""
    
    ws.merge_cells('A4:I10')
    ws['A4'] = about_text
    ws['A4'].font = Font(size=10)
    ws['A4'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    ws['A4'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws['A4'].border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # USING THE TEMPLATES section
    ws['A12'] = 'USING THE TEMPLATES'
    ws['A12'].font = Font(bold=True, size=12)
    ws.row_dimensions[12].height = 20
    
    using_text = """We have created the Processing Activities Registers in Excel for ease of use and filtering and as with all of our documents, you are free to corporate brand, edit and customise the content. It is important to make the fields and entries relevant to your business and sector. We have not used dropdown menus in this register, as many of the fields require bespoke entries. However, you are free to add data validation menus if you have repeated entry requirements that are consistent.

There are 2 registers, one for controllers and one for processors. If you act in the capacity as both, complete both tabs as applicable (i.e. the controller register for your controller processing (e.g. employee records), and the processor register for your processing activities (e.g. CRB checks).

The content of the template has been added to a 'table' which allows for simple filtering and sorting of the columns. This is done by using the arrows in the heading fields. Once you have completed the registers, you can then sort by heading title to access select sections. (See the sample screenshot to the right)"""
    
    ws.merge_cells('A13:I20')
    ws['A13'] = using_text
    ws['A13'].font = Font(size=10)
    ws['A13'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    ws['A13'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws['A13'].border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # EXAMPLES section
    ws['A22'] = 'EXAMPLES'
    ws['A22'].font = Font(bold=True, size=12)
    ws.row_dimensions[22].height = 20
    
    examples_text = """We have added an example completed register on tab 3 which is for guidance only. If should delete this tab once read to avoid mixing up the completed sheets."""
    
    ws.merge_cells('A23:I25')
    ws['A23'] = examples_text
    ws['A23'].font = Font(size=10, italic=True)
    ws['A23'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    ws['A23'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws['A23'].border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Set row heights for text blocks
    for row in range(4, 11):
        ws.row_dimensions[row].height = 25
    for row in range(13, 21):
        ws.row_dimensions[row].height = 25
    for row in range(23, 26):
        ws.row_dimensions[row].height = 25

def generate_populated_ropa_template(export_data_df):
    """Generate ROPA template populated with specific export data"""
    try:
        print(f"Generating populated ROPA template with {len(export_data_df)} records")
        
        # Create workbook
        wb = Workbook()
        
        # Remove the default sheet
        wb.remove(wb.active)

        # Create Introduction sheet first
        create_introduction_sheet(wb)
        
        # Create Controller Processing Activities Register sheet with export data
        create_controller_sheet(wb, export_data_df)
        
        # Create Processor Processing Activity sheet (empty for now)
        create_processor_sheet(wb)

        # Save to temporary file
        temp_dir = tempfile.gettempdir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ROPA_Export_{timestamp}.xlsx"
        file_path = os.path.join(temp_dir, filename)

        wb.save(file_path)
        print(f"Populated ROPA template saved to: {file_path}")

        return file_path

    except Exception as e:
        print(f"Error during populated template generation: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def generate_ropa_template():
    """Generate complete ROPA template with Controller, Processor, and Example sheets"""
    try:
        # First, try to read the uploaded ROPA file to understand its structure
        ropa_file_path = "attached_assets/ROPA_1755785640068.xlsx"
        reference_structure = None
        
        if os.path.exists(ropa_file_path):
            reference_structure = read_uploaded_excel_structure(ropa_file_path)
            if reference_structure is not None:
                print("Successfully read reference ROPA structure")

        # Get existing data to populate template
        existing_data = get_all_ropa_data_for_template()

        # Create workbook
        wb = Workbook()
        
        # Remove the default sheet
        wb.remove(wb.active)

        # Create Introduction sheet first
        create_introduction_sheet(wb)
        
        # Create Controller Processing Activities Register sheet
        create_controller_sheet(wb, existing_data)
        
        # Create Processor Processing Activity sheet
        create_processor_sheet(wb)

        # Save to temporary file
        temp_dir = tempfile.mkdtemp()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ROPA_Complete_{timestamp}.xlsx"
        file_path = os.path.join(temp_dir, filename)

        wb.save(file_path)
        print(f"Complete template saved to: {file_path}")

        return file_path

    except Exception as e:
        print(f"Error during template generation: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
